import requests
import openpyxl
from lxml import etree
url="https://www.gushiwen.cn/"
header={"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36"}
resp=requests.get(url=url,headers=header)

html=etree.HTML(resp.text)
s=' '
for i in range(75):
    xpaths="/html/body/div[2]/div[2]/div[1]/div[2]/a["+str(i)+"]/text()"
    biaoti=html.xpath(xpaths)
    s=s+"\n"+str(biaoti)
print(s)

biao1=openpyxl.Workbook()
sheet=biao1.create_sheet("sheet1")
for i in range(2):
    sheet.cell(1,i+1,s[i])
biao1.save('1.xlsx')
# from ntpath import join
# import re
# import requests
# from lxml import etree
# s=['https://m.dytt8.net/']
# for i in range(1,2):
    
#     url="https://m.dytt8.net/html/gndy/dyzz/list_23_"+str(i)+".html"
#     resp=requests.get(url)
#     resp.encoding='gb2312'
#     html=etree.HTML(resp.text)
#     for x in range(1,25):
#         final=html.xpath("/html//table["+str(x)+"]//tr[2]//b/a/text()")
#         finalweb=html.xpath("/html//table["+str(x)+"]//tr[2]//b/a/@href")
#         # print(final)
       
#         x=s+finalweb
#         w=''.join(x)
#         print(final)
#         print(w)
        
#         # with open('2.xls','a',encoding="utf-8")as ff:
#         #     ff.write(final[0]+'\t')
#         #     ff.write(w+'\n')

        